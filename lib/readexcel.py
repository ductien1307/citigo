import random, string
import openpyxl
import os

currentPath = os.getcwd()
pathImportExcel = currentPath

def openExcelPython(excelName):
    path =  pathImportExcel + excelName
    wb_obj = openpyxl.load_workbook(path)
    excelOpened = wb_obj.active
    return excelOpened

def getAllRowValuePython(excelOpened, rowNum):
    list = []
    m_column = excelOpened.max_column
    rowNum = int(rowNum)
    for i in range(1, m_column + 1):
        cell_obj = excelOpened.cell(row = rowNum, column = i)
        value = cell_obj.value
        list.append(value)
    return list

def getAllColumnValuePython(excelOpened, columnNum):
    list = []
    m_row = excelOpened.max_row
    columnNum = int(columnNum)
    for i in range(1, m_row + 1):
        cell_obj = excelOpened.cell(row = i, column = columnNum)
        value = cell_obj.value
        list.append(value)
    return list

def getCellValuePython(excelOpened, rowNum, columnNum):
    rowNum = int(rowNum)
    columnNum = int(columnNum)
    cell_obj = excelOpened.cell(row = rowNum, column = columnNum)
    value = cell_obj.value
    return value

def getAllColumnValueFromSpecialRowPython(excelOpened, columnNum, rowNum):
    list = []
    m_row = excelOpened.max_row
    columnNum = int(columnNum)
    rowNum = int(rowNum)
    for i in range(rowNum, m_row + 1):
        cell_obj = excelOpened.cell(row = i, column = columnNum)
        value = cell_obj.value
        list.append(value)
    return list
